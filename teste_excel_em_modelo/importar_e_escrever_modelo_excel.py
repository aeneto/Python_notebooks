import pandas as pd
from openpyxl import load_workbook
#importar arquivo para um dataframe:
dados = pd.read_csv(r"C:\Users\André Neto\Documents\cursos\teste_excel_em_modelo\dados.csv",sep=';')

#ordenar coluna competência:
dados.sort_values(by=['competência'])

#contar valores Nan (nulos):
dados.isnull().sum()

#substituir valores nulos por zero:
dados['consulta'].fillna(0,inplace=True)

#carrego o Excel com o template pré-formatado 'template.xlsx':
book = load_workbook(r"C:\Users\André Neto\Documents\cursos\teste_excel_em_modelo\modelo.xlsx")

#defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx':
writer = pd.ExcelWriter(r"C:\Users\André Neto\Documents\cursos\teste_excel_em_modelo\arquivo_editado.xlsx", engine='openpyxl')


#incluo a formatação no writer:
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

#Pegar todos os grupos de contratos existentes na qual será necessário para inserir os dados nas abas:
agrupado = dados['Empresa'].unique()

#For para inserir os dados em todas as abas:
for contrato in agrupado:
    dados_selecionados = dados['Empresa']==contrato
    dados_filtrados = dados[dados_selecionados]
    
    #Criar um data frame com todas as competências
    tabela_1 = ('01/01/2021','01/02/2021','01/03/2021','01/04/2021','01/05/2021','01/06/2021','01/07/2021','01/08/2021','01/09/2021','01/10/2021','01/11/2021','01/12/2021' )
    df = pd.DataFrame(tabela_1,columns=['competência'])

    m = pd.merge(df, dados_filtrados,  on = 'competência', how = 'left')
    dx = pd.melt(df)
    #Seleciona quais colunas serao necessárias do dataframe:
    dados_selecao = m[['consulta','exame','despesa_hospitalar']]

    # para escrever só os valores em um lugar específico. Neste caso, considerou a variável contrato com nome de uma aba:
    dados_selecao.to_excel(writer, sheet_name = str(contrato), startrow=2, startcol=2, header=False, index=False)
    dx.to_excel(writer, sheet_name = str(contrato), startrow=15, startcol=2, header=False, index=False)
    writer.save()
writer.close()

'''
#Seleciona quais colunas serçao necessárias do dataframe
dados_selecao = dados[['consulta','exame','despesa_hospitalar']]

#carrego o Excel com o template pré-formatado 'template.xlsx'
book = load_workbook(r"C:\Users\André Neto\Documents\cursos\teste_excel_em_modelo\modelo.xlsx")

#defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
writer = pd.ExcelWriter(r"C:\Users\André Neto\Documents\cursos\teste_excel_em_modelo\arquivo_editado.xlsx", engine='openpyxl')

#incluo a formatação no writer
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

# para escrever só os valores em um lugar específico:
dados_selecao.to_excel(writer, 'Planilha1', startrow=2, startcol=2, header=False, index=False)

'''
